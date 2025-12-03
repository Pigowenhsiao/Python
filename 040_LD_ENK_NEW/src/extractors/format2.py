from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl as px
try:
    from openpyxl.worksheet.read_only import ReadOnlyWorksheet
except ImportError:
    ReadOnlyWorksheet = px.worksheet.worksheet.Worksheet  # type: ignore
import xlrd

from ..lib import SQL

from ..common_log import log_error, log_info

from ..config_loader import PipelineConfig, MappingEntry


@dataclass
class SheetAccessor:
    mode: str  # xls or xlsx
    sheet: object

    def get_value(self, cell: str):
        if self.mode == "xlsx":
            return self.sheet[cell].value
        row, col = _cell_to_indices(cell)
        return self.sheet.cell(row, col).value


def _cell_to_indices(cell: str):
    column = 0
    row_part = ""
    for char in cell:
        if char.isalpha():
            column = column * 26 + (ord(char.upper()) - 64)
        else:
            row_part += char
    row = int(row_part)
    return row - 1, column - 1


class Format2Extractor:
    def __init__(self, config: PipelineConfig, log_file: str):
        self.config = config
        self.log_file = log_file
        if not config.mapping_groups:
            raise ValueError("Format2 requires at least one mapping group")
        self.mapping = next(iter(config.mapping_groups.values()))
        self.serial_initials = self._load_serial_initials()

    def extract(self, workbook_path: str) -> List[Dict[str, object]]:
        log_info(self.log_file, f"Opening mapper workbook {workbook_path}")
        path = Path(workbook_path)
        ext = path.suffix.lower()
        if ext == ".xls":
            return self._extract_xls(path)
        return self._extract_xlsx(path)

    def _extract_xls(self, path: Path) -> List[Dict[str, object]]:
        log_info(self.log_file, "Using xlrd path (.xls detected)")
        wb = xlrd.open_workbook(path, on_demand=True)
        sheet = self._get_xls_sheet(wb, self.config.excel.sheet_name)
        date_sheet_name = self.config.excel.date_sheet_name or self.config.excel.sheet_name
        date_sheet = self._get_xls_sheet(wb, date_sheet_name)
        if sheet is None or date_sheet is None:
            wb.release_resources()
            log_error(self.log_file, f"Workbook {path} missing required sheet(s)")
            return []
        accessor = SheetAccessor("xls", sheet)
        records = self._build_record(accessor, date_sheet)
        wb.release_resources()
        return [records] if records else []

    def _extract_xlsx(self, path: Path) -> List[Dict[str, object]]:
        log_info(self.log_file, "Using openpyxl path (.xlsx/.xlsm detected)")
        wb = px.load_workbook(path, data_only=True, read_only=True)
        sheet = self._get_xlsx_sheet(wb, self.config.excel.sheet_name)
        date_sheet_name = self.config.excel.date_sheet_name or self.config.excel.sheet_name
        date_sheet = self._get_xlsx_sheet(wb, date_sheet_name)
        if sheet is None or date_sheet is None:
            wb.close()
            log_error(self.log_file, f"Workbook {path} missing required sheet(s)")
            return []
        accessor = SheetAccessor("xlsx", sheet)
        record = self._build_record(accessor, date_sheet)
        wb.close()
        if record:
            log_info(self.log_file, "Sheet extracted successfully")
            return [record]
        log_error(self.log_file, "Record extraction returned empty")
        return []

    def _build_record(self, accessor: SheetAccessor, date_sheet) -> Optional[Dict[str, object]]:
        serial_cell = accessor.get_value("G3")
        if not serial_cell:
            return None
        serial_number = str(serial_cell).strip()

        if self.serial_initials and not any(
            serial_number.startswith(token) for token in self.serial_initials
        ):
            log_info(self.log_file, f"Serial {serial_number} skipped (not in whitelist)")
            return None

        conn, cursor = SQL.connSQL()
        if conn is None:
            log_error(self.log_file, f"{serial_number} : Connection with Prime Failed")
            return None
        try:
            part_number, lot_number = SQL.selectSQL(cursor, serial_number)
        finally:
            SQL.disconnSQL(conn, cursor)

        data: Dict[str, object] = {
            "key_serial_number": serial_number,
            "key_part_number": part_number,
            "key_LotNumber_9": lot_number,
        }
        operator = accessor.get_value("E3")
        data["key_operator"] = "-" if not operator else operator
        start_date = self._read_date(date_sheet)
        if start_date:
            data["key_start_date_time"] = start_date

        for entry in self.mapping.entries:
            value = self._read_entry(entry, accessor)
            data[entry.key] = value

        data["Operation"] = self.config.general.operation
        data["TestStation"] = self.config.general.test_station
        data["Site"] = self.config.general.site
        return data

    def _read_date(self, date_sheet) -> Optional[str]:
        date_cell = self.config.excel.date_cell or "A5"
        if isinstance(date_sheet, px.worksheet.worksheet.Worksheet):
            raw = date_sheet[date_cell].value
        elif isinstance(date_sheet, ReadOnlyWorksheet):
            row, col = _cell_to_indices(date_cell)
            raw = date_sheet.cell(row=row + 1, column=col + 1).value
        else:
            row, col = _cell_to_indices(date_cell)
            raw = date_sheet.cell(row + 1, col + 1).value
        if raw is None:
            return None
        return str(raw).replace(" ", "T")

    def _read_entry(self, entry: MappingEntry, accessor: SheetAccessor):
        if entry.aggregate:
            values = []
            for ref in entry.references:
                cell = ref.split("!", 1)[1] if "!" in ref else ref
                raw = accessor.get_value(cell)
                if raw in (None, ""):
                    continue
                try:
                    values.append(float(raw))
                except (TypeError, ValueError):
                    continue
            if not values:
                return ""
            if entry.aggregate == "AVG":
                return sum(values) / len(values)
            return ""
        ref = entry.cell
        return accessor.get_value(ref)

    def _load_serial_initials(self) -> Optional[List[str]]:
        path = self.config.paths.serial_whitelist
        if not path or not path.exists():
            return None
        with path.open("r", encoding="utf-8") as fh:
            return [line.strip() for line in fh if line.strip()]

    def _get_xlsx_sheet(self, workbook, sheet_name: Optional[str]):
        if sheet_name:
            try:
                return workbook[sheet_name]
            except KeyError:
                log_error(
                    self.log_file,
                    f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}",
                )
                return None
        first = workbook.sheetnames[0]
        log_info(self.log_file, f"No sheet name specified, using first sheet '{first}'")
        return workbook[first]

    def _get_xls_sheet(self, workbook, sheet_name: Optional[str]):
        if sheet_name:
            try:
                return workbook.sheet_by_name(sheet_name)
            except xlrd.biffh.XLRDError:
                log_error(
                    self.log_file,
                    f"Sheet '{sheet_name}' not found. Available: {workbook.sheet_names()}",
                )
                return None
        log_info(self.log_file, "No sheet name specified, using first sheet (index 0)")
        return workbook.sheet_by_index(0)
