# Pythonプログラム - すべての.iniファイルを読み取り、データ処理を実行してXMLファイルを生成
import os
import sys
import glob
import shutil
import logging
import pandas as pd
from configparser import ConfigParser, NoSectionError, NoOptionError
from datetime import datetime, timedelta

# カスタムモジュールのインポート
sys.path.append('../MyModule')
import Log, SQL, Check, Convert_Date, Row_Number_Func
from openpyxl import load_workbook
import random
import logging

# グローバル変数: ログファイル
global_log_file = None

# ログ設定の構成
def setup_logging(log_file_path):
    try:
        logging.basicConfig(filename=log_file_path, level=logging.DEBUG,
                            format='%(asctime)s - %(levelname)s - %(message)s',
                            datefmt='%Y-%m-%d %H:%M:%S')
    except OSError as e:
        Log.Log_Error(global_log_file, f"Error setting up logging: {e}")
        raise

# 実行記録ファイルを更新する関数
def update_running_rec(running_rec_path, end_date):
    try:
        with open(running_rec_path, 'w', encoding='utf-8') as f:
            end_date = end_date.replace('.', ':').replace('T', ' ')
            end_date_dt = datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S')
            f.write(end_date_dt.strftime('%Y-%m-%d %H:%M:%S'))
        Log.Log_Info(global_log_file, f"実行記録ファイル {running_rec_path} を終了日 {end_date} で更新しました")
    except Exception as e:
        Log.Log_Error(global_log_file, f"実行記録ファイル {running_rec_path} の更新エラー: {e}")

# XMLファイルを生成する関数
def generate_xml(data_dict):
    try:
        start_date_time = data_dict.get('key_Start_Date_Time', '')
        if start_date_time:
            dt = datetime.strptime(start_date_time, '%Y-%m-%dT%H.%M.%S')
            random_seconds = random.randint(0, 59)
            dt = dt.replace(second=random_seconds)
            data_dict['key_Start_Date_Time'] = dt.strftime('%Y-%m-%dT%H.%M.%S')

        xml_filename = f"Site={site},ProductFamily={product_family},Operation={operation},PartNumber={data_dict.get('key_Part_Number', 'Unknown')},SerialNumber={data_dict.get('key_Serial_Number', 'Unknown')},Testdate={data_dict.get('key_Start_Date_Time', 'Unknown')}.xml"
        xml_filepath = os.path.join(output_path, xml_filename)

        with open(xml_filepath, 'w', encoding='utf-8') as f:
            f.write('<?xml version="1.0" encoding="utf-8"?>\n')
            f.write('<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">\n')
            f.write(f'    <Result startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Result="Passed">\n')
            f.write(f'        <Header SerialNumber="{data_dict["key_Serial_Number"]}" PartNumber="{data_dict["key_Part_Number"]}" Operation="{operation}" TestStation="{Test_Station}" Operator="{data_dict["key_Operator"]}" StartTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Site="{site}" LotNumber="{data_dict["key_Serial_Number"]}"/>\n')
            f.write('        <HeaderMisc>\n')
            f.write(f'            <Item Description="{operation}"/>\n')
            f.write('        </HeaderMisc>\n')
            f.write(f'        <TestStep Name="{data_dict["key_Operation"]}" startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Status="Passed">\n')
            f.write(f'            <Data DataType="String" Name="Banchi_ID" Value="{data_dict["key_Banchi_ID"]}"/>\n')
            f.write(f'            <Data DataType="Numeric" Name="Current" Units="uA" Value="{data_dict["key_Current"]}"/>\n')
            f.write(f'            <Data DataType="Numeric" Name="Voltage" Units="V" Value="{data_dict["key_Voltage"]}"/>\n')
            f.write(f'            <Data DataType="Numeric" Name="{data_dict["key_Banchi_ID"]}" Units="uA" Value="{data_dict["key_Current"]}"/>\n')
            f.write('        </TestStep>\n')
            f.write(f'        <TestStep Name="SORTED_DATA" startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Status="Passed">\n')
            f.write(f'            <Data DataType="Numeric" Name="STARTTIME_SORTED" Value="{data_dict["key_STARTTIME_SORTED"]}"/>\n')
            f.write(f'            <Data DataType="String" Name="LotNumber_5" Value="{data_dict["key_Serial_Number"]}" CompOperation="LOG"/>\n')
            f.write(f'            <Data DataType="String" Name="LotNumber_9" Value="{data_dict["key_LotNumber_9"]}" CompOperation="LOG"/>\n')
            f.write('        </TestStep>\n')
            f.write('        <TestEquipment>\n')
            f.write(f'            <Item DeviceName="MOCVD" DeviceSerialNumber="{data_dict["Tool_ID"]}"/>\n')
            f.write('        </TestEquipment>\n')
            f.write('    </Result>\n')
            f.write('</Results>\n')

        Log.Log_Info(global_log_file, f'XML File Created: {xml_filepath}')
    except Exception as e:
        Log.Log_Error(global_log_file, f"Failed to create XML file for SerialNumber={data_dict.get('key_Serial_Number', 'Unknown')}: {e}")

# Excelファイルを処理する関数
def process_excel_file(file_path):
    workbook = load_workbook(file_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        Log.Log_Error(global_log_file, f"Sheet '{sheet_name}' not found in the workbook. Skipping file: {file_path}")
        return

    sheet = workbook[sheet_name]
    tool_id_value = sheet[Tool_ID].value
    Log.Log_Info(global_log_file, f"Extracted Tool_ID value: {tool_id_value}")
    try:
        tool_id_value = tool_id_value.split('-')[0]
    except AttributeError:
        tool_id_value = "No_Tool_data"
    print(tool_id_value)
    try:
        # Excelデータを読み取る
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, usecols=data_columns, skiprows=Title_Row, nrows=int(Data_Row))
        df_result = pd.DataFrame()
        df_result = df.loc[[int(Title_Row)-Title_Row, int(Data_Row)-Title_Row-1]]
        df_result.reset_index(drop=True, inplace=True)
        complete_df = pd.DataFrame()
        # NaN値を含む列を削除
        df_result = df_result.dropna(axis=1, how='all')
        # インデックスをリセット
        df_result = df_result.reset_index(drop=True)
        df_result.columns = range(df_result.shape[1])  # 列インデックスをリセット
        file_mod_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
        for i in range(0, int(len(df_result.columns)/4)):
            cell_value = df_result.loc[0, 4 * i]
            parts = cell_value.split('_')
            if len(parts) == 2:
                serial_number, banchi_loc = parts
                banchi_id, loc = banchi_loc.split('-')
            else:
                serial_number = cell_value
                banchi_id = loc = ''
            new_df = pd.DataFrame({
                        'Serial_Number': [serial_number],
                        'Banchi-ID': [banchi_id],
                        'Loc': [loc],
                        'Volt': [df_result.loc[1, 4 * i]],
                        'Current': [max(df_result.loc[1, 4 * i + 1], df_result.loc[1, 4 * i + 3])],
                        'Start_date_time': [datetime.strptime(file_mod_time, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%dT%H.%M.%S')]
                    })
            complete_df = pd.concat([complete_df, new_df], ignore_index=True)
        complete_df = complete_df.sort_values(by=['Serial_Number', 'Banchi-ID', 'Current'], ascending=[True, True, False]).drop_duplicates(subset=['Serial_Number', 'Banchi-ID'], keep='first')
    except Exception as e:
        Log.Log_Error(global_log_file, f'Error reading Excel file {file_path}: {e}')
        Log.Log_Error(global_log_file, f'Ensure the file exists, is not corrupted, and the parameters (sheet_name, usecols, skiprows, nrows) are correct.')
        return
    if 'Serial_Number' not in complete_df.columns:
        Log.Log_Error(global_log_file, f"'Serial_Number' column missing in the Excel file {file_path}. Check the file structure.")
        return
    complete_df['Part_Number'] = None  # 'Part_Number'列を確保
    Serial_Number = complete_df['Serial_Number'].tolist()

    conn, cursor = SQL.connSQL()
    if conn is None:
        Log.Log_Error(global_log_file, 'Connection with Prime Failed')
        return
    try:
        for serial in Serial_Number:
            part_number, nine_serial_number = SQL.selectSQL(cursor, serial)
            if part_number and nine_serial_number:
                complete_df.loc[complete_df['Serial_Number'] == serial, 'Part_Number'] = part_number
                complete_df.loc[complete_df['Serial_Number'] == serial, 'Nine_Serial_Number'] = nine_serial_number
            else:
                Log.Log_Error(global_log_file, f'Serial number {serial} not found in database')
   
    except Exception as e:
        Log.Log_Error(global_log_file, f'SQL query failed: {e}')
    finally:
        SQL.disconnSQL(conn, cursor)
    
    try:
        with open(running_rec, 'a', encoding='utf-8') as f:
            print(f"{file_path}\n")
            f.write(f"{file_path}\n")
        Log.Log_Info(global_log_file, f"Appended processed file path to running_rec: {file_path}")
    except Exception as e:
        Log.Log_Error(global_log_file, f"Error appending file path to running_rec: {e}")
    
        # 'Part_Number'がNaNの行を削除
    complete_df = complete_df.dropna(subset=['Part_Number'])
        # 列数をリセット
    complete_df = complete_df.reset_index(drop=True)
    row_number = 0        
    Log.Log_Info(global_log_file, f'Processing dataframe {len(complete_df)} rows')
    # 成功したファイルのパスと名前をBanchi-IV_StartRow.txtに追加
    try:
        with open('./Banchi-IV_StartRow.txt', 'a', encoding='utf-8') as f:
            f.write(f"{file_path}\n")
        Log.Log_Info(global_log_file, f"Appended processed file path to Banchi-IV_StartRow.txt: {file_path}")
    except Exception as e:
        Log.Log_Error(global_log_file, f"Error appending file path to Banchi-IV_StartRow.txt: {e}")

        # データ処理
    for row_number in range(len(complete_df)):
        data_dict = {}
            # データ変換処理
            # 最新のkey_Start_Date_Timeで実行記録を更新
        
        latest_date = complete_df['Start_date_time'].max()
        #update_running_rec(running_rec, latest_date)
        data_dict["key_Start_Date_Time"]=latest_date                
        data_dict['key_Operation'] = operation
        date = datetime.strptime(str(data_dict["key_Start_Date_Time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])
        data_dict['key_Serial_Number'] = complete_df.loc[row_number, 'Serial_Number']
        data_dict['key_Operator'] = 'Unknown'
        data_dict['key_Banchi_ID'] = complete_df.loc[row_number, 'Banchi-ID']
        data_dict['key_Voltage'] = complete_df.loc[row_number, 'Volt']
        data_dict['key_Current'] = complete_df.loc[row_number, 'Current']
        data_dict['Tool_ID']=tool_id_value
            
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        if complete_df.loc[row_number, 'Part_Number'] is not None:
            data_dict['key_Part_Number'] = complete_df.loc[row_number, 'Part_Number']
            data_dict['key_LotNumber_9'] = complete_df.loc[row_number, 'Nine_Serial_Number']
        else:
            Log.Log_Error(global_log_file, data_dict.get('key_Serial_Number', 'Unknown') + ' : ' + 'PartNumber Error')
            row_number += 1
            continue
            # XMLファイルを生成
        if None in data_dict.values():
            Log.Log_Error(global_log_file, f"Skipping row {row_number} due to None values in data_dict")
        else:
            generate_xml(data_dict)


def process_ini_file(config_path):
    global global_log_file
    config = ConfigParser()
    try:
        with open(config_path, 'r', encoding='utf-8') as config_file:
            config.read_file(line for line in config_file if not line.strip().startswith('#'))
    except Exception as e:
        Log.Log_Error(global_log_file, f"Error reading config file {config_path}: {e}")
        return

    # 設定ファイルから設定を取得
    try:
        global input_paths, output_path, running_rec, sheet_name, data_columns, log_path, site, product_family, operation, Test_Station, file_name_pattern, exclude_dirs, Title_Row, Data_Row,Tool_ID
        input_paths = [path.strip() for path in config.get('Paths', 'input_paths').split(',')]
        output_path = config.get('Paths', 'output_path')
        running_rec = config.get('Paths', 'running_rec')
        sheet_name = config.get('Excel', 'sheet_name')
        data_columns = config.get('Excel', 'data_columns')
        log_path = config.get('Logging', 'log_path')
        site = config.get('Basic_info', 'Site')
        product_family = config.get('Basic_info', 'ProductFamily')
        operation = config.get('Basic_info', 'Operation')
        Test_Station = config.get('Basic_info', 'TestStation')
        file_name_pattern = config.get('Basic_info', 'file_name_pattern')
        exclude_dirs = config.get('Basic_info', 'exclude_dirs').split(',')
        Title_Row = config.getint('Excel', 'Title_Row')
        Data_Row = config.getint('Excel', 'Data_Row')
        Tool_ID = config.get('Excel','Tool')
    except NoSectionError as e:
        Log.Log_Error(global_log_file, f"Missing section in config file {config_path}: {e}")
        return
    except NoOptionError as e:
        Log.Log_Error(global_log_file, f"Missing option in config file {config_path}: {e}")
        return

    # ログフォルダとファイルを作成
    log_folder_name = str(datetime.today().date())
    log_folder_path = os.path.join(log_path, log_folder_name)
    if not os.path.exists(log_folder_path):
        os.makedirs(log_folder_path)
    log_file = os.path.join(log_folder_path, '043_LD-SPUT.log')
    global_log_file = log_file

    # ログ設定を行う
    setup_logging(global_log_file)
    Log.Log_Info(log_file, f'Program Start for config {config_path}')
    
    #file_name_pattern='*.xlsx'
    Log.Log_Info(log_file, 'Searching Banchi IV file')
    for input_path in input_paths:
        for root, dirs, files in os.walk(input_path):
            #print(root)
            dirs[:] = [d for d in dirs if not d[0].isdigit() and d not in exclude_dirs]
            for file in files:
                if file.startswith('$'):  # Skip files starting with '$'
                    continue
                if glob.fnmatch.fnmatch(file, file_name_pattern):
                    file_path = os.path.join(root, file)
                    file_mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                    if (datetime.now() - file_mod_time).days <= 10:  # Setting data retrieval date
                        Log.Log_Info(log_file, f'Processing file {file_path}')
                        process_excel_file(file_path)

                
# すべての.iniファイルをスキャンして処理するメイン関数
def main():
    ini_files = glob.glob("*.ini")
    for ini_file in ini_files:
        process_ini_file(ini_file)

if __name__ == '__main__':
    main()

Log.Log_Info(global_log_file, 'Program End')
