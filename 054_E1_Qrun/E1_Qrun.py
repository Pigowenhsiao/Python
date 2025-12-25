"""
E1_Qrun uploader: read INI, parse Excel, compute stats, output CSV/XML, optional DB lookup and dedup.
スクリプトはINI読み込み、Excel解析、統計計算、CSV/XML出力、DB照会と重複除去（任意）を行う。
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import sqlite3
import sys
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from configparser import ConfigParser

try:
    import openpyxl  # type: ignore
except Exception as exc:  # pragma: no cover
    raise RuntimeError("openpyxl is required to read .xlsx/.xlsm") from exc


# -------------------------
# Legacy module path & import (align with sample code) / レガシーモジュールのパスと読み込み（サンプルに合わせる）
# -------------------------
# Align with your legacy structure: sys.path.append('../MyModule') / 既存構成に合わせて sys.path を追加
# This script assumes it's located under a project folder where ../MyModule exists. / 本スクリプトは ../MyModule が存在する構成を前提とする
sys.path.append(str((Path(__file__).resolve().parent / "../MyModule").resolve()))

try:
    import SQL  # type: ignore
except Exception as exc:  # pragma: no cover
    raise RuntimeError(
        "Failed to import legacy SQL module. Ensure ../MyModule/SQL.py exists and is importable."
    ) from exc


# -------------------------
# Constants / Regex / 定数・正規表現
# -------------------------
# Accept filenames like: / 例として許可されるファイル名
#   N3250317_N3025先行結果.xlsm
#   N1250922_N1004先行結果.xlsx
# Exclude ones containing "- Copy" by logic (not regex) / "- Copy" を含むものはロジックで除外
FILENAME_RE = re.compile(r"^(N[A-Z0-9]{7})_(N[A-Z0-9]{4}).*?\.(xlsx|xlsm)$", re.IGNORECASE)

STAT_SUFFIXES = ("MAX", "MIN", "AVG", "STD")


# =========================
# Dataclasses: Settings / データクラス（設定）
# =========================

@dataclass(frozen=True)
class BasicInfo:
    output_mode: str
    site: str
    product_family: str
    operation: str
    test_station: str
    file_name_patterns: List[str]
    retention_date_days: int
    tool_name_fallback: str
    part_number_default: str


@dataclass(frozen=True)
class PathsSettings:
    input_paths: List[Path]
    running_rec: Path
    output_path: Path
    csv_path: Path
    intermediate_data_path: Path
    log_path: Path


@dataclass(frozen=True)
class ExcelSettings:
    sheet_name: str
    data_columns: str
    main_skip_rows: int
    main_nrows: int


@dataclass(frozen=True)
class StartDateTimeSettings:
    sheet_name: str
    cell: str
    datetime_format: str
    fallback_mode: str  # file_mtime / now / blank
    output_format: str


@dataclass(frozen=True)
class DedupSettings:
    enable_dedup: bool
    skip_dedup_check: bool
    db_path: Path
    fingerprint_mode: str  # stat / sha256
    debug_discovery: bool
    debug_limit_10_files: bool
    db_filter_by_mtime: bool
    db_filter_days: int


@dataclass(frozen=True)
class WaiveLengthSettings:
    mapping: Dict[str, str]  # LotRule -> Waive_Leng_Cate
    missing_rule_behavior: str  # unknown / skip_file / error
    unknown_value: str


@dataclass(frozen=True)
class DatabaseSettings:
    # Keep all INI fields (even if legacy SQL module doesn't use them directly) / INI項目は未使用でも保持
    db_connection_string: str
    server: str
    database: str
    username: str
    password: str
    driver: str

    # Output masking policy for CSV (default masked) / CSV 出力のマスク方針（既定はマスク）
    password_mask: str = "***"


@dataclass(frozen=True)
class DataField:
    key: str
    col: str
    dtype: str

    def is_assigned_by_python(self) -> bool:
        return self.col.strip() == "-1"

    def is_excel_col_index(self) -> bool:
        c = self.col.strip()
        return c.isdigit() or (c.startswith("-") and c[1:].isdigit())

    def excel_index(self) -> int:
        return int(self.col.strip())

    def is_cell_ref(self) -> bool:
        return self.col.strip().lower().startswith("cell_")


@dataclass(frozen=True)
class Settings:
    basic: BasicInfo
    paths: PathsSettings
    excel: ExcelSettings
    start_dt: StartDateTimeSettings
    dedup: DedupSettings
    waive: WaiveLengthSettings
    db: DatabaseSettings
    fields: List[DataField]


# =========================
# Logging / ロギング
# =========================

def setup_logging(log_dir: Path, operation: str) -> logging.Logger:
    """
    Set up rotating loggers for file and console output.
    ファイルとコンソール出力用のローテーションログを設定する。
    """
    date_folder = log_dir / datetime.now().strftime("%Y-%m-%d")
    date_folder.mkdir(parents=True, exist_ok=True)
    log_file = date_folder / f"{operation}_{datetime.now().strftime('%Y%m%d')}.log"

    logger = logging.getLogger(operation)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(name)s | %(message)s"
    )

    file_handler = RotatingFileHandler(
        filename=str(log_file),
        maxBytes=10 * 1024 * 1024,  # 10MB
        backupCount=20,
        encoding="utf-8",
    )
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


# =========================
# Config parsing / 設定読み込み
# =========================

def _read_ini(path: Path) -> ConfigParser:
    """
    Read INI using UTF-8 (with BOM) and no interpolation.
    UTF-8（BOM可）でINIを読み込み、補間を無効化する。
    """
    cfg = ConfigParser(interpolation=None)
    cfg.read(path, encoding="utf-8")
    return cfg


def _get_list(cfg: ConfigParser, section: str, key: str) -> List[str]:
    raw = cfg.get(section, key, fallback="").strip()
    if not raw:
        return []
    parts = [line.strip() for line in raw.splitlines() if line.strip()]
    if len(parts) == 1 and "," in parts[0]:
        return [p.strip() for p in parts[0].split(",") if p.strip()]
    return parts


def _parse_fields(cfg: ConfigParser) -> List[DataField]:
    raw = cfg.get("DataFields", "fields", fallback="")
    fields: List[DataField] = []
    for line in raw.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or line.startswith(";"):
            continue
        if ":" not in line:
            continue
        parts = [p.strip() for p in line.split(":", 2)]
        if len(parts) != 3:
            continue
        fields.append(DataField(key=parts[0], col=parts[1], dtype=parts[2]))
    return fields


def _parse_waive_mapping(cfg: ConfigParser) -> Dict[str, str]:
    if not cfg.has_section("WaiveLengthCategoryMapping"):
        return {}
    items = dict(cfg.items("WaiveLengthCategoryMapping"))
    return {k.strip().upper(): v.strip() for k, v in items.items()}


def load_settings(ini_path: Path) -> Settings:
    """
    Parse INI into Settings dataclasses.
    INIを読み取りSettingsデータクラスに変換する。
    """
    cfg = _read_ini(ini_path)

    basic = BasicInfo(
        output_mode=cfg.get("Basic_info", "output_mode", fallback="csv").strip(),
        site=cfg.get("Basic_info", "Site", fallback="").strip(),
        product_family=cfg.get("Basic_info", "ProductFamily", fallback="").strip(),
        operation=cfg.get("Basic_info", "Operation", fallback="").strip(),
        test_station=cfg.get("Basic_info", "TestStation", fallback="").strip(),
        file_name_patterns=_get_list(cfg, "Basic_info", "file_name_patterns"),
        retention_date_days=cfg.getint("Basic_info", "Retention_date", fallback=7),
        tool_name_fallback=cfg.get("Basic_info", "Tool_Name", fallback="").strip(),
        part_number_default=cfg.get("Basic_info", "Part_Number", fallback="HL13E1").strip(),
    )

    input_paths = _get_list(cfg, "Paths", "input_paths")
    paths = PathsSettings(
        input_paths=[Path(p) for p in input_paths],
        running_rec=Path(cfg.get("Paths", "running_rec", fallback="./running.txt")),
        output_path=Path(cfg.get("Paths", "output_path", fallback="./output_xml")),
        csv_path=Path(cfg.get("Paths", "CSV_path", fallback="./output_csv")),
        intermediate_data_path=Path(cfg.get("Paths", "intermediate_data_path", fallback="./intermediate")),
        log_path=Path(cfg.get("Paths", "log_path", fallback="./log")),
    )

    excel = ExcelSettings(
        sheet_name=cfg.get("Excel", "sheet_name", fallback="").strip(),
        data_columns=cfg.get("Excel", "data_columns", fallback="").strip(),
        main_skip_rows=cfg.getint("Excel", "main_skip_rows", fallback=0),
        main_nrows=cfg.getint("Excel", "main_nrows", fallback=0),
    )

    start_dt = StartDateTimeSettings(
        sheet_name=cfg.get("StartDateTime", "sheet_name", fallback="").strip(),
        cell=cfg.get("StartDateTime", "cell", fallback="").strip(),
        datetime_format=cfg.get("StartDateTime", "datetime_format", fallback="").strip(),
        fallback_mode=cfg.get("StartDateTime", "fallback_mode", fallback="file_mtime").strip().lower(),
        output_format=cfg.get("StartDateTime", "output_format", fallback="%Y-%m-%d %H:%M:%S").strip(),
    )

    dedup = DedupSettings(
        enable_dedup=cfg.getboolean("Dedup", "enable_dedup", fallback=True),
        skip_dedup_check=cfg.getboolean("Dedup", "skip_dedup_check", fallback=False),
        db_path=Path(cfg.get("Dedup", "dedup_db_path", fallback="./dedup_registry.sqlite")),
        fingerprint_mode=cfg.get("Dedup", "fingerprint_mode", fallback="stat").strip().lower(),
        debug_discovery=cfg.getboolean("Dedup", "debug_discovery", fallback=False),
        debug_limit_10_files=cfg.getboolean("Dedup", "debug_limit_10_files", fallback=False),
        db_filter_by_mtime=cfg.getboolean("Dedup", "db_filter_by_mtime", fallback=False),
        db_filter_days=cfg.getint("Dedup", "db_filter_days", fallback=30),
    )

    waive = WaiveLengthSettings(
        mapping=_parse_waive_mapping(cfg),
        missing_rule_behavior=cfg.get("WaiveLengthCategory", "missing_rule_behavior", fallback="unknown").strip().lower(),
        unknown_value=cfg.get("WaiveLengthCategory", "unknown_value", fallback="UNKNOWN").strip(),
    )

    db = DatabaseSettings(
        db_connection_string=cfg.get("Database", "db_connection_string", fallback="").strip(),
        server=cfg.get("Database", "server", fallback="").strip(),
        database=cfg.get("Database", "database", fallback="").strip(),
        username=cfg.get("Database", "username", fallback="").strip(),
        password=cfg.get("Database", "password", fallback="").strip(),
        driver=cfg.get("Database", "driver", fallback="").strip(),
    )

    fields = _parse_fields(cfg)

    return Settings(
        basic=basic,
        paths=paths,
        excel=excel,
        start_dt=start_dt,
        dedup=dedup,
        waive=waive,
        db=db,
        fields=fields,
    )


# =========================
# Dedup registry (SQLite) / 重複排除レジストリ（SQLite）
# =========================

class DedupRegistry:
    """
    SQLite-backed registry for processed files.
    処理済みファイルを管理するSQLiteレジストリ。
    """

    def __init__(self, db_path: Path, logger: logging.Logger) -> None:
        self.db_path = db_path
        self.logger = logger
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()

    def _init_db(self) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS processed_files (
                    fingerprint TEXT PRIMARY KEY,
                    file_path TEXT NOT NULL,
                    file_size INTEGER NOT NULL,
                    file_mtime REAL NOT NULL,
                    processed_at TEXT NOT NULL,
                    status TEXT NOT NULL,
                    output_csv TEXT,
                    output_xml TEXT,
                    message TEXT
                )
                """
            )
            conn.commit()

    def has(self, fingerprint: str) -> bool:
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.execute(
                "SELECT 1 FROM processed_files WHERE fingerprint = ? LIMIT 1",
                (fingerprint,),
            )
            return cur.fetchone() is not None

    def add(
        self,
        fingerprint: str,
        file_path: Path,
        file_size: int,
        file_mtime: float,
        status: str,
        output_csv: Optional[Path],
        output_xml: Optional[Path],
        message: str = "",
    ) -> None:
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO processed_files
                (fingerprint, file_path, file_size, file_mtime, processed_at, status, output_csv, output_xml, message)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    fingerprint,
                    str(file_path),
                    int(file_size),
                    float(file_mtime),
                    datetime.now().isoformat(timespec="seconds"),
                    status,
                    str(output_csv) if output_csv else None,
                    str(output_xml) if output_xml else None,
                    message,
                ),
            )
            conn.commit()


def compute_fingerprint(path: Path, mode: str) -> str:
    """
    Compute a fingerprint for a file (stat or sha256).
    ファイルの指紋を算出する（stat/sha256）。
    """
    st = path.stat()
    if mode == "sha256":
        h = hashlib.sha256()
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()

    raw = f"{path.resolve()}|{st.st_size}|{st.st_mtime}".encode("utf-8", errors="ignore")
    return hashlib.sha256(raw).hexdigest()


# =========================
# Excel helpers / Excel ヘルパー
# =========================

def _read_cell(workbook_path: Path, sheet_name: str, cell: str) -> Any:
    """
    Read one cell from an Excel sheet.
    Excelシートから単一セルを読む。
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    value = ws[cell].value
    wb.close()
    return value


def _coerce_datetime(value: Any, fmt_hint: str) -> Optional[datetime]:
    """
    Coerce a value into datetime using format hints.
    形式ヒントを使って値をdatetimeに変換する。
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return base + pd.to_timedelta(float(value), unit="D")  # type: ignore[arg-type]
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            if fmt_hint:
                return datetime.strptime(text, fmt_hint)
            parsed = pd.to_datetime(text, errors="coerce")
            if pd.isna(parsed):
                return None
            return parsed.to_pydatetime()
        except Exception:
            return None
    return None


def read_start_datetime(
    file_path: Path,
    settings: StartDateTimeSettings,
    logger: logging.Logger,
) -> str:
    """
    Read Start_Date_Time from sheet/cell with fallback.
    シート/セルからStart_Date_Timeを取得し、フォールバックする。
    """
    dt_obj: Optional[datetime] = None
    try:
        if settings.sheet_name and settings.cell:
            value = _read_cell(file_path, settings.sheet_name, settings.cell)
            dt_obj = _coerce_datetime(value, settings.datetime_format)
    except Exception as exc:
        logger.warning(
            "Failed to read StartDateTime from %s!%s: %s",
            settings.sheet_name,
            settings.cell,
            exc,
        )

    if dt_obj is None:
        if settings.fallback_mode == "now":
            dt_obj = datetime.now()
        elif settings.fallback_mode == "blank":
            return ""
        else:
            dt_obj = datetime.fromtimestamp(file_path.stat().st_mtime)

    return dt_obj.strftime(settings.output_format)


def read_tester_id_ay23(
    file_path: Path,
    main_sheet_name: str,
    logger: logging.Logger,
) -> str:
    """
    Read TESTER_ID from AY23 in the main sheet.
    主シートAY23からTESTER_IDを読む。
    """
    try:
        value = _read_cell(file_path, main_sheet_name, "AY23")
        if value is None:
            return ""
        return str(value).strip()
    except Exception as exc:
        logger.warning("Failed to read TESTER_ID from %s!AY23: %s", main_sheet_name, exc)
        return ""


def read_main_table(
    file_path: Path,
    excel: ExcelSettings,
    logger: logging.Logger,
) -> pd.DataFrame:
    """
    Read the main table range into a DataFrame.
    主テーブル範囲をDataFrameとして読み込む。
    """
    try:
        t0 = datetime.now()
        logger.info("Excel read start: %s", file_path.name)
        df = pd.read_excel(
            file_path,
            sheet_name=excel.sheet_name,
            usecols=excel.data_columns,
            skiprows=excel.main_skip_rows,
            nrows=excel.main_nrows,
            header=None,
            engine="openpyxl",
        )
        logger.info("Excel read end: %s (elapsed=%s)", file_path.name, datetime.now() - t0)
        logger.info("Loaded main table: %s rows x %s cols", df.shape[0], df.shape[1])
        return df
    except Exception as exc:
        raise RuntimeError(f"Failed to read main table from {file_path.name}: {exc}") from exc


# =========================
# Waive length category / Waive 長さカテゴリ
# =========================

def compute_waive_leng_cate(
    lot_id: str,
    waive: WaiveLengthSettings,
    logger: logging.Logger,
) -> str:
    """
    Compute Waive_Leng_Cate from Lot ID prefix.
    Lot ID先頭からWaive_Leng_Cateを算出する。
    """
    lot_rule = lot_id[:2].upper()
    value = waive.mapping.get(lot_rule)

    if value:
        return value

    behavior = waive.missing_rule_behavior
    if behavior == "skip_file":
        raise ValueError(f"LotRule '{lot_rule}' not found (skip_file)")
    if behavior == "error":
        raise KeyError(f"LotRule '{lot_rule}' not found (error)")

    logger.warning("LotRule '%s' not found in mapping; use %s", lot_rule, waive.unknown_value)
    return waive.unknown_value


# =========================
# DB query (align with legacy SQL module) / DB クエリ（レガシーSQLに合わせる）
# =========================

def query_database_via_legacy_sql(
    lot_id: str,
    logger: logging.Logger,
) -> Dict[str, Any]:
    """
    Query DB via legacy SQL module and map fields.
    レガシーSQLモジュールでDB照会しフィールドにマップする。
    """
    conn = None
    cursor = None
    try:
        t0 = datetime.now()
        logger.info("DB query start: LotID=%s", lot_id)
        conn, cursor = SQL.connSQL()
        if conn is None or cursor is None:
            logger.error("Legacy SQL.connSQL() returned None (connection failed).")
            return {}

        result = SQL.selectSQL(cursor, str(lot_id))

        # Normalize result to tuple/list / 戻り値を tuple/list に正規化
        if result is None:
            logger.warning("Legacy SQL.selectSQL() returned None for LotID=%s", lot_id)
            return {}

        if isinstance(result, (list, tuple)):
            values = list(result)
        else:
            # Sometimes it might return a single scalar / 単一値が返る場合に備える
            values = [result]

        # === Default mapping (adjust if your legacy SQL returns different columns) ===
        # 既存SQLの戻り列に合わせて調整する
        out: Dict[str, Any] = {}
        if len(values) >= 3:
            # Expected: LotID, Part_Number, LotNumber_9 / 想定順序: LotID, Part_Number, LotNumber_9
            out["DB_LOOKUP_Part_Number"] = values[1]
            out["LotNumber_9"] = values[2]
        else:
            if len(values) >= 1:
                out["DB_LOOKUP_Part_Number"] = values[0]
            if len(values) >= 2:
                out["LotNumber_9"] = values[1]

        # If there are more fields, append generic names / 追加列は汎用名で追加
        start_idx = 3 if len(values) >= 3 else 2
        for i in range(start_idx, len(values)):
            out[f"DB_LOOKUP_Field_{i+1}"] = values[i]

        logger.info("DB query end: LotID=%s (elapsed=%s)", lot_id, datetime.now() - t0)
        return out

    except Exception as exc:
        logger.exception("DB query failed via legacy SQL for LotID=%s: %s", lot_id, exc)
        return {}
    finally:
        if "t0" in locals():
            logger.info("DB query cleanup: LotID=%s (elapsed=%s)", lot_id, datetime.now() - t0)
        try:
            if conn is not None:
                SQL.disconnSQL(conn, cursor)
        except Exception:
            # Don't crash on disconnect / 切断失敗は無視
            pass


def query_database_bulk_via_legacy_sql(
    lot_ids: Sequence[str],
    logger: logging.Logger,
) -> Dict[str, Dict[str, Any]]:
    """
    Query DB for multiple Lot IDs using one connection.
    1接続で複数Lot IDを照会する。
    """
    if not lot_ids:
        return {}

    t0 = datetime.now()
    logger.info("DB bulk query start: count=%d", len(lot_ids))
    conn = None
    cursor = None
    out: Dict[str, Dict[str, Any]] = {}
    try:
        conn, cursor = SQL.connSQL()
        if conn is None or cursor is None:
            logger.error("Legacy SQL.connSQL() returned None (connection failed).")
            return {}

        for lot_id in lot_ids:
            try:
                result = SQL.selectSQL(cursor, str(lot_id))
                if result is None:
                    logger.warning("Legacy SQL.selectSQL() returned None for LotID=%s", lot_id)
                    out[str(lot_id)] = {}
                    continue

                if isinstance(result, (list, tuple)):
                    values = list(result)
                else:
                    values = [result]

                mapped: Dict[str, Any] = {}
                if len(values) >= 3:
                    mapped["DB_LOOKUP_Part_Number"] = values[1]
                    mapped["LotNumber_9"] = values[2]
                    start_idx = 3
                else:
                    if len(values) >= 1:
                        mapped["DB_LOOKUP_Part_Number"] = values[0]
                    if len(values) >= 2:
                        mapped["LotNumber_9"] = values[1]
                    start_idx = 2

                for i in range(start_idx, len(values)):
                    mapped[f"DB_LOOKUP_Field_{i+1}"] = values[i]

                out[str(lot_id)] = mapped
            except Exception as exc:
                logger.exception("DB query failed via legacy SQL for LotID=%s: %s", lot_id, exc)
                out[str(lot_id)] = {}

        logger.info("DB bulk query end: count=%d (elapsed=%s)", len(lot_ids), datetime.now() - t0)
        return out
    finally:
        logger.info("DB bulk query cleanup: count=%d (elapsed=%s)", len(lot_ids), datetime.now() - t0)
        try:
            if conn is not None:
                SQL.disconnSQL(conn, cursor)
        except Exception:
            pass


# =========================
# Stats computation / 統計計算
# =========================

def _stat_name_from_key(key: str) -> str:
    """
    Convert DataFields key_* to the base name.
    DataFieldsのkey_*を基底名に変換する。
    """
    if key.startswith("key_"):
        return key[4:]
    return key


def compute_stats_for_fields(
    df: pd.DataFrame,
    fields: Sequence[DataField],
    logger: logging.Logger,
) -> Dict[str, Any]:
    """
    Compute MAX/MIN/AVG/STD for measurement fields.
    計測フィールドのMAX/MIN/AVG/STDを計算する。
    """
    out: Dict[str, Any] = {}

    meas = [f for f in fields if f.is_excel_col_index() and f.excel_index() >= 0]
    for f in meas:
        base = _stat_name_from_key(f.key)
        idx = f.excel_index()
        if idx >= df.shape[1]:
            logger.warning("Field %s col=%s out of range (df cols=%s)", f.key, idx, df.shape[1])
            for stat in STAT_SUFFIXES:
                out[f"{base}_{stat}"] = None
            continue

        series = pd.to_numeric(df.iloc[:, idx], errors="coerce")
        out[f"{base}_MAX"] = series.max(skipna=True)
        out[f"{base}_MIN"] = series.min(skipna=True)
        out[f"{base}_AVG"] = series.mean(skipna=True)
        out[f"{base}_STD"] = series.std(skipna=True, ddof=1)

    return out


# =========================
# CSV / XML output / CSV・XML 出力
# =========================

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def make_output_csv_path(csv_dir: Path, operation: str) -> Path:
    """
    Build a unique CSV output path.
    一意なCSV出力パスを生成する。
    """
    ensure_dir(csv_dir)
    ts = datetime.now().strftime("%Y_%m_%dT%H.%M.%S")
    short = uuid.uuid4().hex[:8]
    return csv_dir / f"{operation}_{ts}_{short}.csv"


def append_csv(csv_path: Path, row: Mapping[str, Any]) -> None:
    """
    Append one row to CSV, adding header if needed.
    1行をCSVに追記し、必要ならヘッダを付ける。
    """
    df = pd.DataFrame([dict(row)])
    exists = csv_path.exists()
    df.to_csv(csv_path, mode="a", header=not exists, index=False, encoding="utf-8-sig")


def generate_pointer_xml(
    output_dir: Path,
    settings: Settings,
    csv_path: Path,
    serial_no: str,
) -> Path:
    """
    Generate pointer XML referencing the CSV.
    CSV参照のポインタXMLを生成する。
    """
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    ensure_dir(output_dir)
    now_iso = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    xml_file_name = (
        f"Site={settings.basic.site},"
        f"ProductFamily={settings.basic.product_family},"
        f"Operation={settings.basic.operation},"
        f"Partnumber={settings.basic.part_number_default},"
        f"Serialnumber={serial_no},"
        f"Testdate={now_iso}.xml"
    ).replace(":", ".")

    xml_path = output_dir / xml_file_name

    results = ET.Element(
        "Results",
        {
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xmlns:xsd": "http://www.w3.org/2001/XMLSchema",
        },
    )
    result = ET.SubElement(
        results,
        "Result",
        startDateTime=now_iso,
        endDateTime=now_iso,
        Result="Passed",
    )
    ET.SubElement(
        result,
        "Header",
        SerialNumber=serial_no,
        PartNumber=settings.basic.part_number_default,
        Operation=settings.basic.operation,
        TestStation=settings.basic.test_station,
        Operator="NA",
        StartTime=now_iso,
        Site=settings.basic.site,
        LotNumber="",
    )
    test_step = ET.SubElement(
        result,
        "TestStep",
        Name=settings.basic.operation,
        startDateTime=now_iso,
        endDateTime=now_iso,
        Status="Passed",
    )
    ET.SubElement(
        test_step,
        "Data",
        DataType="Table",
        Name=f"tbl_{settings.basic.operation.upper()}",
        Value=str(csv_path),
        CompOperation="LOG",
    )

    xml_bytes = minidom.parseString(ET.tostring(results)).toprettyxml(
        indent="  ", encoding="utf-8"
    )
    with xml_path.open("wb") as f:
        f.write(xml_bytes)

    return xml_path


# =========================
# File discovery / ファイル探索
# =========================

def discover_source_files(
    input_paths: Sequence[Path],
    patterns: Sequence[str],
    logger: logging.Logger,
    debug: bool = False,
) -> List[Path]:
    """
    Find source files matching patterns in input paths.
    入力パスでパターン一致ファイルを探索する。
    """
    found: List[Path] = []

    if debug:
        print()
        print("[DEBUG] ===== Discover Source Files =====")

    for base in input_paths:
        if debug:
            print(f"[DEBUG] Scan path: {base}")

        if not base.exists():
            if debug:
                print(f"[DEBUG] ❌ Path not exists: {base}")
            logger.warning("Input path not found: %s", base)
            continue

        if debug:
            all_files = list(base.iterdir())
            print(f"[DEBUG]   Total files in dir: {len(all_files)}")
            for p in all_files:
                print(f"[DEBUG]     - {p.name}")

        for pat in patterns:
            if debug:
                print(f"[DEBUG]   Try pattern: {pat}")
            matched = list(base.glob(pat))
            if debug:
                print(f"[DEBUG]     Matched count: {len(matched)}")

            for p in matched:
                if p.name.startswith("~$"):
                    if debug:
                        print(f"[DEBUG]       Skip temp file: {p.name}")
                    continue

                if debug:
                    print(f"[DEBUG]       ✔ Matched: {p.name}")
                found.append(p)

    uniq = sorted({p.resolve() for p in found})

    if debug:
        print(f"[DEBUG] ===== Result: {len(uniq)} candidate files =====")
        print()
    logger.info("Discovered %d candidate files.", len(uniq))

    return uniq


def parse_filename_ids(file_path: Path) -> Tuple[str, str]:
    """
    Parse wafer_id and lot_id from filename.
    Exclude names containing "- Copy" and allow suffixes like "先行結果".
    ファイル名からwafer_idとlot_idを抽出する。
    "- Copy" を含むものは除外し、「先行結果」等の後継を許可する。
    """
    name = file_path.name

    if "- Copy" in name or "copy" in name.lower():
        raise ValueError(f"Excluded copied file: {name}")

    m = FILENAME_RE.match(name)
    if not m:
        raise ValueError(f"Invalid filename format: {name}")

    wafer_id = m.group(1).upper()
    lot_id = m.group(2).upper()
    return wafer_id, lot_id


# =========================
# Per-file processing / ファイル単位の処理
# =========================

def build_output_row(
    file_path: Path,
    settings: Settings,
    logger: logging.Logger,
    db_lookup_map: Optional[Mapping[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """
    Build the final output row for one wafer file.
    One output row per file (summary row).
    1ファイルにつき1行（サマリー行）を出力する。
    """
    _wafer_id, lot_id = parse_filename_ids(file_path)

    serial_number = lot_id
    start_dt_str = read_start_datetime(file_path, settings.start_dt, logger)

    waive_leng = compute_waive_leng_cate(lot_id, settings.waive, logger)

    # Default Part_Number comes from INI; override when DB provides a value / 既定はINIの値、DB値があれば上書き
    part_number = settings.basic.part_number_default

    tester_id = read_tester_id_ay23(file_path, settings.excel.sheet_name, logger)
    if not tester_id:
        tester_id = settings.basic.tool_name_fallback

    # DB query using legacy module (bulk map preferred) / レガシーSQLでDB参照
    if db_lookup_map is not None:
        db_lookup_fields = db_lookup_map.get(lot_id, {})
    else:
        db_lookup_fields = query_database_via_legacy_sql(lot_id, logger)
    db_part_number = db_lookup_fields.pop("DB_LOOKUP_Part_Number", None)
    if db_part_number is not None and str(db_part_number).strip():
        part_number = str(db_part_number).strip()
    db_lotnumber_9 = db_lookup_fields.get("LotNumber_9")

    # Read main data and compute stats / 主データ読込と統計計算
    df = read_main_table(file_path, settings.excel, logger)
    stats = compute_stats_for_fields(df, settings.fields, logger)

    # DB connection info columns removed from CSV by request / DB接続情報はCSVに出力しない
    row: Dict[str, Any] = {
        "Serial_Number": serial_number,
        "Start_Date_Time": start_dt_str,
        "Part_Number": part_number,
        "TESTER_ID": tester_id,
        "Waive_Leng_Cate": waive_leng,
        "LotNumber_9": db_lotnumber_9 if db_lotnumber_9 is not None else "",
        "Source_File": file_path.name,
    }

    # Merge DB lookup results / DB検索結果を結合
    row.update(db_lookup_fields)

    # Merge stats / 統計結果を結合
    row.update(stats)
    return row


# =========================
# Program entry / エントリポイント
# =========================

def run_for_ini(ini_path: Path) -> None:
    """
    Run processing for one INI file.
    1つのINIに対して処理を実行する。
    """
    settings = load_settings(ini_path)
    logger = setup_logging(settings.paths.log_path, settings.basic.operation)

    logger.info("==== Start E1_Qrun uploader ====")
    logger.info("INI: %s", ini_path.resolve())

    ensure_dir(settings.paths.csv_path)
    ensure_dir(settings.paths.output_path)
    ensure_dir(settings.paths.intermediate_data_path)

    csv_out = make_output_csv_path(settings.paths.csv_path, settings.basic.operation)
    logger.info("CSV output: %s", csv_out)

    registry: Optional[DedupRegistry] = None
    if settings.dedup.enable_dedup and not settings.dedup.skip_dedup_check:
        registry = DedupRegistry(settings.dedup.db_path, logger)
        logger.info("Dedup enabled: %s", settings.dedup.db_path)
    else:
        logger.info(
            "Dedup skipped (enable_dedup=%s, skip_dedup_check=%s)",
            settings.dedup.enable_dedup,
            settings.dedup.skip_dedup_check,
        )

    files = discover_source_files(
        settings.paths.input_paths,
        settings.basic.file_name_patterns,
        logger,
        debug=settings.dedup.debug_discovery,
    )

    if settings.dedup.debug_limit_10_files:
        files = files[:10]
        logger.info("Debug limit enabled: processing only %d files.", len(files))

    # Pre-collect Lot IDs for one-shot DB querying
    lot_ids: List[str] = []
    cutoff_dt: Optional[datetime] = None
    if settings.dedup.db_filter_by_mtime:
        cutoff_dt = datetime.now() - timedelta(days=settings.dedup.db_filter_days)
    for f in files:
        try:
            _wafer_id, lot_id = parse_filename_ids(f)
            if cutoff_dt is not None:
                mtime = datetime.fromtimestamp(f.stat().st_mtime)
                if mtime < cutoff_dt:
                    continue
            lot_ids.append(lot_id)
        except Exception:
            continue

    db_lookup_map: Optional[Dict[str, Dict[str, Any]]] = None
    if lot_ids:
        if cutoff_dt is not None:
            logger.info(
                "DB filter enabled: keep files with mtime >= %s (%d days).",
                cutoff_dt.strftime("%Y-%m-%d %H:%M:%S"),
                settings.dedup.db_filter_days,
            )
        db_lookup_map = query_database_bulk_via_legacy_sql(lot_ids, logger)

    processed_count = 0
    skipped_count = 0
    error_count = 0

    for f in files:
        try:
            # strict filename check (also excludes "- Copy") / 厳格なファイル名チェック（"- Copy" も除外）
            parse_filename_ids(f)

            fp = compute_fingerprint(f, settings.dedup.fingerprint_mode)
            st = f.stat()

            if registry and registry.has(fp):
                skipped_count += 1
                logger.info("Skip (dedup): %s", f.name)
                continue

            row = build_output_row(f, settings, logger, db_lookup_map=db_lookup_map)
            append_csv(csv_out, row)

            processed_count += 1
            logger.info("Processed: %s", f.name)

            if registry:
                registry.add(
                    fingerprint=fp,
                    file_path=f,
                    file_size=int(st.st_size),
                    file_mtime=float(st.st_mtime),
                    status="SUCCESS",
                    output_csv=csv_out,
                    output_xml=None,
                    message="",
                )

        except Exception as exc:
            error_count += 1
            logger.exception("Failed processing file %s: %s", f.name, exc)
            if not f.exists():
                continue
            try:
                fp = compute_fingerprint(f, settings.dedup.fingerprint_mode)
                st = f.stat()
                if registry:
                    registry.add(
                        fingerprint=fp,
                        file_path=f,
                        file_size=int(st.st_size),
                        file_mtime=float(st.st_mtime),
                        status="FAILED",
                        output_csv=csv_out if csv_out.exists() else None,
                        output_xml=None,
                        message=str(exc),
                    )
            except Exception:
                pass

    xml_out: Optional[Path] = None
    if csv_out.exists() and csv_out.stat().st_size > 0:
        serial_for_xml = csv_out.stem
        xml_out = generate_pointer_xml(settings.paths.output_path, settings, csv_out, serial_for_xml)
        logger.info("Pointer XML generated: %s", xml_out)

    logger.info("==== End ====")
    logger.info(
        "Processed=%d Skipped=%d Errors=%d CSV=%s XML=%s",
        processed_count,
        skipped_count,
        error_count,
        csv_out,
        xml_out,
    )


def main(argv: Sequence[str]) -> int:
    """
    CLI entry point.
    CLIエントリポイント。
    """
    os.chdir(Path(__file__).resolve().parent)

    if len(argv) >= 2:
        ini_path = Path(argv[1])
        if not ini_path.exists():
            print(f"INI not found: {ini_path}")
            return 2
        run_for_ini(ini_path)
        return 0

    ini_files = sorted(Path(".").glob("*.ini"))
    if not ini_files:
        print("No .ini files found in current directory.")
        return 1

    for ini in ini_files:
        run_for_ini(ini)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
